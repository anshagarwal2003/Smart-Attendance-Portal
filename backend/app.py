from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory, send_file
from werkzeug.utils import secure_filename

import sqlite3
import random
import string
import base64
import os
import math
import io
import threading
import time
from datetime import datetime, timedelta

from huggingface_hub import snapshot_download, HfApi

import numpy as np
from PIL import Image, ImageOps
import face_recognition

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Force IST Timezone on Linux servers (Hugging Face)
if hasattr(time, 'tzset'):
    os.environ['TZ'] = 'Asia/Kolkata'
    time.tzset()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)

HF_TOKEN = os.environ.get("HF_TOKEN")
DATASET_REPO_ID = "anshagarwal2003/smart-attendance-data"

if HF_TOKEN:
    try:
        print("Downloading persistent data from HF Dataset...")
        snapshot_download(
            repo_id=DATASET_REPO_ID,
            repo_type="dataset",
            local_dir=DATA_DIR,
            token=HF_TOKEN
        )
    except Exception as e:
        print(f"Dataset empty or failed to download: {e}")

    def sync_to_hf():
        api = HfApi()
        while True:
            time.sleep(300) # Sync every 5 minutes
            try:
                api.upload_folder(
                    folder_path=DATA_DIR,
                    repo_id=DATASET_REPO_ID,
                    repo_type="dataset",
                    token=HF_TOKEN
                )
                print("Successfully synced data to HF Dataset.")
            except Exception as e:
                print(f"Failed to sync data: {e}")

    # Start background sync thread
    sync_thread = threading.Thread(target=sync_to_hf, daemon=True)
    sync_thread.start()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "smart-attendance-secret")
app.config.update(
    SESSION_COOKIE_SAMESITE='None',
    SESSION_COOKIE_SECURE=True
)

DB_NAME = os.environ.get("DATABASE_PATH", os.path.join(DATA_DIR, "smart_attendance.db"))

ATTENDANCE_WINDOW_MINUTES = 15

COLLEGE_LAT = 26.9124
COLLEGE_LNG = 75.7873
ALLOWED_RADIUS_METERS = 100000

CLASSROOM_LAT = 26.9124
CLASSROOM_LNG = 75.7873
TEACHER_ALLOWED_RADIUS_METERS = 100000

TEACHER_ACTIVATION_LIMIT_MINUTES = 15
FACE_MATCH_THRESHOLD = 0.50

PROOF_FOLDER = os.environ.get("PROOF_FOLDER", os.path.join(DATA_DIR, "proof_images"))
STUDENT_IMAGE_FOLDER = os.environ.get("STUDENT_IMAGE_FOLDER", os.path.join(DATA_DIR, "student_images"))
EXPORT_FOLDER = os.environ.get("EXPORT_FOLDER", os.path.join(DATA_DIR, "exports"))

for folder in [PROOF_FOLDER, STUDENT_IMAGE_FOLDER, EXPORT_FOLDER]:
    os.makedirs(folder, exist_ok=True)


def get_db():
    con = sqlite3.connect(DB_NAME)
    con.row_factory = sqlite3.Row
    return con


def normalize_section(section):
    return str(section).strip().upper().replace(" ", "-")


def get_course_from_section(section):
    section = normalize_section(section)

    if "-" in section:
        return section.split("-")[0]

    return section


def get_all_sections_for_course(con, course):
    course = course.strip().upper()
    sections_set = set()

    explicit = con.execute("SELECT section_name FROM sections WHERE UPPER(course_name) = ?", (course,)).fetchall()
    for row in explicit:
        sec = str(row["section_name"]).strip().upper()
        if not sec.startswith(course + "-"):
            sec = f"{course}-{sec}"
        sections_set.add(sec)

    like_pattern = course + "-%"
    students_secs = con.execute("""
        SELECT DISTINCT section FROM students 
        WHERE UPPER(REPLACE(TRIM(section), ' ', '-')) LIKE ? OR UPPER(TRIM(section)) = ?
    """, (like_pattern, course)).fetchall()
    for row in students_secs:
        sections_set.add(normalize_section(row["section"]))

    timetable_secs = con.execute("""
        SELECT DISTINCT section FROM timetable 
        WHERE UPPER(REPLACE(TRIM(section), ' ', '-')) LIKE ? OR UPPER(TRIM(section)) = ?
    """, (like_pattern, course)).fetchall()
    for row in timetable_secs:
        sections_set.add(normalize_section(row["section"]))

    if not sections_set:
        sections_set.add(f"{course}-A")

    return sorted(list(sections_set))


def get_all_sections(con):
    sections_set = set()
    rows = con.execute("""
        SELECT course_name, section_name FROM sections
    """).fetchall()
    for row in rows:
        c = row["course_name"].strip().upper()
        s = row["section_name"].strip().upper()
        if not s.startswith(c + "-"):
            sections_set.add(f"{c}-{s}")
        else:
            sections_set.add(s)

    for table in ["students", "timetable"]:
        rows = con.execute(f"SELECT DISTINCT section FROM {table}").fetchall()
        for row in rows:
            if row["section"]:
                sections_set.add(normalize_section(row["section"]))

    return sorted(list(sections_set))


def generate_code(length=6):
    letters = string.ascii_uppercase + string.digits
    return "".join(random.choice(letters) for _ in range(length))


def close_expired_active_sessions():
    con = get_db()
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    con.execute("""
        UPDATE class_sessions
        SET status = 'CLOSED'
        WHERE class_date <= ?
        AND status = 'ACTIVE'
        AND attendance_end < ?
    """, (current_date, current_time))

    con.commit()
    con.close()


def calculate_distance(lat1, lng1, lat2, lng2):
    radius = 6371000

    p1 = math.radians(lat1)
    p2 = math.radians(lat2)

    delta_p = math.radians(lat2 - lat1)
    delta_l = math.radians(lng2 - lng1)

    a = (
        math.sin(delta_p / 2) ** 2
        + math.cos(p1) * math.cos(p2) * math.sin(delta_l / 2) ** 2
    )

    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius * c


def base64_to_rgb_array(image_data):
    if "," in image_data:
        image_data = image_data.split(",", 1)[1]

    image_bytes = base64.b64decode(image_data)

    image = Image.open(io.BytesIO(image_bytes))
    image = ImageOps.exif_transpose(image)
    image = image.convert("RGB")

    image_array = np.array(image, dtype=np.uint8)
    image_array = np.ascontiguousarray(image_array)

    return image_array


def get_student_image_path(roll_no):
    roll_no = str(roll_no).strip().upper()
    possible_extensions = [".jpg", ".jpeg", ".png", ".webp"]

    for ext in possible_extensions:
        path = os.path.join(STUDENT_IMAGE_FOLDER, roll_no + ext)
        if os.path.exists(path):
            return path

    return None


def verify_student_face(image_data, roll_no):
    registered_image_path = get_student_image_path(roll_no)

    if registered_image_path is None:
        return False, "Registered student photo not found in student_images folder"

    try:
        known_pil = Image.open(registered_image_path)
        known_pil = ImageOps.exif_transpose(known_pil)
        known_pil = known_pil.convert("RGB")

        known_image = np.array(known_pil, dtype=np.uint8)
        known_image = np.ascontiguousarray(known_image)

        captured_image = base64_to_rgb_array(image_data)

        known_encodings = face_recognition.face_encodings(known_image)

        if len(known_encodings) == 0:
            return False, "No face found in registered student photo"

        captured_encodings = face_recognition.face_encodings(captured_image)

        if len(captured_encodings) == 0:
            return False, "No face found in captured camera image"

        known_encoding = known_encodings[0]
        captured_encoding = captured_encodings[0]

        face_distance = face_recognition.face_distance(
            [known_encoding],
            captured_encoding
        )[0]

        if face_distance <= FACE_MATCH_THRESHOLD:
            return True, f"Face matched successfully. Distance: {face_distance:.3f}"

        return False, f"Face not matched. Distance: {face_distance:.3f}"

    except Exception as e:
        return False, f"Face verification error: {str(e)}"


def save_proof_image(image_data, roll_no, session_id):
    if "," in image_data:
        image_data = image_data.split(",", 1)[1]

    image_bytes = base64.b64decode(image_data)

    filename = f"{roll_no}_session_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
    filepath = os.path.join(PROOF_FOLDER, filename)

    with open(filepath, "wb") as file:
        file.write(image_bytes)

    return filename


@app.route("/proof_images/<path:filename>")
def proof_image(filename):
    return send_from_directory(PROOF_FOLDER, filename)


def init_db():
    con = get_db()
    cur = con.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS admins (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        admin_id TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS teachers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        teacher_id TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        roll_no TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        section TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS timetable (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        day TEXT NOT NULL,
        subject TEXT NOT NULL,
        start_time TEXT NOT NULL,
        end_time TEXT NOT NULL,
        section TEXT NOT NULL,
        teacher_id INTEGER NOT NULL,
        room_no TEXT,
        room_range INTEGER
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS rooms (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        room_no TEXT NOT NULL UNIQUE,
        room_range INTEGER NOT NULL,
        room_lat REAL,
        room_lng REAL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS class_sessions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id INTEGER NOT NULL,
        subject TEXT NOT NULL,
        section TEXT NOT NULL,
        class_date TEXT NOT NULL,
        class_time TEXT NOT NULL,
        activated_at TEXT NOT NULL,
        attendance_end TEXT NOT NULL,
        session_code TEXT NOT NULL,
        status TEXT NOT NULL,
        timetable_id INTEGER
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id INTEGER NOT NULL,
        student_id INTEGER NOT NULL,
        date TEXT NOT NULL,
        time TEXT,
        status TEXT NOT NULL,
        distance_meters REAL,
        proof_image TEXT,
        UNIQUE(session_id, student_id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS courses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        course_name TEXT UNIQUE NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS sections (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        course_name TEXT NOT NULL,
        section_name TEXT NOT NULL,
        UNIQUE(course_name, section_name)
    )
    """)

    con.commit()

    # Automatically migrate missing columns in existing SQLite databases
    migrations = [
        ("timetable", "room_no", "TEXT"),
        ("timetable", "room_range", "INTEGER"),
        ("rooms", "room_lat", "REAL"),
        ("rooms", "room_lng", "REAL"),
        ("class_sessions", "timetable_id", "INTEGER"),
        ("attendance", "distance_meters", "REAL"),
        ("attendance", "proof_image", "TEXT"),
    ]
    for table_name, col_name, col_type in migrations:
        try:
            cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_type}")
        except Exception:
            pass
    con.commit()

    admin_count = cur.execute("SELECT COUNT(*) AS total FROM admins").fetchone()["total"]
    if admin_count == 0:
        cur.execute(
            "INSERT INTO admins(admin_id, password) VALUES (?, ?)",
            ("admin", "admin123")
        )

    teacher_count = cur.execute("SELECT COUNT(*) AS total FROM teachers").fetchone()["total"]
    if teacher_count == 0:
        cur.execute(
            "INSERT INTO teachers(name, teacher_id, password) VALUES (?, ?, ?)",
            ("Demo Teacher", "T101", "1234")
        )

    student_count = cur.execute("SELECT COUNT(*) AS total FROM students").fetchone()["total"]
    if student_count == 0:
        students = [
            ("Ansh Agarwal", "23BCON2214", "1234", "CSE-A"),
            ("Harshita Mangal", "23BCON0235", "1234", "CSE-A"),
            ("Darshan Singh", "23BCON0142", "1234", "CSE-A"),
        ]

        cur.executemany(
            "INSERT INTO students(name, roll_no, password, section) VALUES (?, ?, ?, ?)",
            students
        )

    timetable_count = cur.execute("SELECT COUNT(*) AS total FROM timetable").fetchone()["total"]
    if timetable_count == 0:
        teacher = cur.execute(
            "SELECT id FROM teachers WHERE teacher_id = ?",
            ("T101",)
        ).fetchone()

        teacher_db_id = teacher["id"]

        timetable = [
            ("Monday", "JAVA", "08:00", "09:00", "CSE-A", teacher_db_id),
            ("Tuesday", "CC", "08:00", "09:00", "CSE-A", teacher_db_id),
            ("Wednesday", "PYTHON", "08:00", "09:00", "CSE-A", teacher_db_id),
            ("Thursday", "DBMS", "08:00", "09:00", "CSE-A", teacher_db_id),
            ("Friday", "MATHS", "08:00", "09:00", "CSE-A", teacher_db_id),
            ("Saturday", "JAVA LAB", "08:00", "09:00", "CSE-A", teacher_db_id),
        ]

        cur.executemany("""
            INSERT INTO timetable(day, subject, start_time, end_time, section, teacher_id)
            VALUES (?, ?, ?, ?, ?, ?)
        """, timetable)

    con.commit()
    con.close()


@app.route("/")
def home():
    return render_template("index.html")

@app.route("/server-time")
def server_time():
    return jsonify({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "day": datetime.now().strftime("%A"),
        "tz": os.environ.get("TZ", "Not Set")
    })


@app.route("/teacher-login", methods=["GET", "POST"])
def teacher_login():
    if request.method == "POST":
        teacher_id = request.form.get("teacher_id", "").strip()
        password = request.form.get("password", "").strip()

        con = get_db()
        teacher = con.execute(
            "SELECT * FROM teachers WHERE teacher_id = ? AND password = ?",
            (teacher_id, password)
        ).fetchone()
        con.close()

        if teacher:
            session.clear()
            session["teacher_db_id"] = teacher["id"]
            session["teacher_name"] = teacher["name"]
            return redirect(url_for("teacher_dashboard"))

        return render_template("teacher_login.html", error="Wrong Teacher ID or Password")

    return render_template("teacher_login.html")


@app.route("/student-login", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        roll_no = request.form.get("roll_no", "").strip().upper()
        password = request.form.get("password", "").strip()

        con = get_db()
        student = con.execute(
            "SELECT * FROM students WHERE UPPER(roll_no) = ? AND password = ?",
            (roll_no, password)
        ).fetchone()
        con.close()

        if student:
            session.clear()
            session["student_db_id"] = student["id"]
            session["student_name"] = student["name"]
            session["roll_no"] = student["roll_no"]
            session["section"] = student["section"]
            return redirect(url_for("student_dashboard"))

        return render_template("student_login.html", error="Wrong Roll No or Password")

    return render_template("student_login.html")


@app.route("/teacher-dashboard")
def teacher_dashboard():
    if "teacher_db_id" not in session:
        return redirect(url_for("teacher_login"))

    close_expired_active_sessions()

    today = datetime.now().strftime("%A")
    current_date = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now()

    con = get_db()

    sections = con.execute("""
        SELECT 
            section,
            COUNT(*) AS total_classes
        FROM timetable
        WHERE teacher_id = ? AND day = ?
        GROUP BY section
        ORDER BY section
    """, (session["teacher_db_id"], today)).fetchall()

    active_session = con.execute("""
        SELECT * FROM class_sessions
        WHERE teacher_id = ? 
        AND class_date = ?
        AND status = 'ACTIVE'
        ORDER BY id DESC LIMIT 1
    """, (
        session["teacher_db_id"],
        current_date
    )).fetchone()

    today_classes_raw = con.execute("""
        SELECT *
        FROM timetable
        WHERE teacher_id = ?
        AND day = ?
        ORDER BY start_time
    """, (session["teacher_db_id"], today)).fetchall()

    upcoming_classes = []
    can_start_classes = []

    for cls in today_classes_raw:
        class_start = datetime.strptime(
            current_date + " " + cls["start_time"],
            "%Y-%m-%d %H:%M"
        )
        
        class_end = datetime.strptime(
            current_date + " " + cls["end_time"],
            "%Y-%m-%d %H:%M"
        )

        can_start_window = class_start - timedelta(minutes=15)
        class_time = f"{cls['start_time']} - {cls['end_time']}"

        existing_session = con.execute("""
            SELECT *
            FROM class_sessions
            WHERE teacher_id = ?
            AND UPPER(subject) = ?
            AND UPPER(REPLACE(TRIM(section), ' ', '-')) = ?
            AND class_date = ?
            AND class_time = ?
            ORDER BY id DESC LIMIT 1
        """, (
            session["teacher_db_id"],
            cls["subject"].upper(),
            normalize_section(cls["section"]),
            current_date,
            class_time
        )).fetchone()

        if existing_session:
            continue

        cls_dict = dict(cls)
        cls_dict["activation_deadline"] = class_end.strftime("%H:%M")

        if now < can_start_window:
            cls_dict["status_type"] = "UPCOMING"
            cls_dict["status_text"] = f"Upcoming at {cls['start_time']}"
            upcoming_classes.append(cls_dict)

        elif can_start_window <= now <= class_end:
            cls_dict["status_type"] = "CAN_START"
            cls_dict["status_text"] = f"Can start till {class_end.strftime('%H:%M')}"
            can_start_classes.append(cls_dict)

    con.close()

    return render_template(
        "teacher_dashboard.html",
        today=today,
        sections=sections,
        active_session=active_session,
        upcoming_classes=upcoming_classes,
        can_start_classes=can_start_classes
    )


@app.route("/teacher-section/<section>")
def teacher_section(section):
    if "teacher_db_id" not in session:
        return redirect(url_for("teacher_login"))

    close_expired_active_sessions()

    section = normalize_section(section)
    today = datetime.now().strftime("%A")
    current_date = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now()

    con = get_db()

    subjects_raw = con.execute("""
        SELECT *
        FROM timetable
        WHERE teacher_id = ?
        AND day = ?
        AND UPPER(REPLACE(TRIM(section), ' ', '-')) = ?
        ORDER BY start_time
    """, (
        session["teacher_db_id"],
        today,
        section
    )).fetchall()

    subjects = []

    for cls in subjects_raw:
        class_start = datetime.strptime(
            current_date + " " + cls["start_time"],
            "%Y-%m-%d %H:%M"
        )

        class_end = datetime.strptime(
            current_date + " " + cls["end_time"],
            "%Y-%m-%d %H:%M"
        )

        activation_deadline = class_start + timedelta(minutes=TEACHER_ACTIVATION_LIMIT_MINUTES)
        class_time = f"{cls['start_time']} - {cls['end_time']}"

        existing_session = con.execute("""
            SELECT *
            FROM class_sessions
            WHERE teacher_id = ?
            AND UPPER(subject) = ?
            AND UPPER(REPLACE(TRIM(section), ' ', '-')) = ?
            AND class_date = ?
            AND class_time = ?
            ORDER BY id DESC LIMIT 1
        """, (
            session["teacher_db_id"],
            cls["subject"].upper(),
            section,
            current_date,
            class_time
        )).fetchone()

        cls_dict = dict(cls)
        cls_dict["activation_deadline"] = activation_deadline.strftime("%H:%M")
        cls_dict["session_id"] = None

        if existing_session:
            cls_dict["session_id"] = existing_session["id"]

            if existing_session["status"] == "ACTIVE":
                cls_dict["status_type"] = "ACTIVE"
                cls_dict["status_text"] = "Running"

            elif existing_session["status"] == "CLOSED":
                cls_dict["status_type"] = "CLASS_OVER"
                cls_dict["status_text"] = "Class Over"

            elif existing_session["status"] == "NOT_HELD":
                cls_dict["status_type"] = "NOT_HELD"
                cls_dict["status_text"] = "Class Not Held"

            else:
                cls_dict["status_type"] = "CLASS_OVER"
                cls_dict["status_text"] = existing_session["status"]

        else:
            if now < class_start:
                cls_dict["status_type"] = "UPCOMING"
                cls_dict["status_text"] = f"Upcoming at {cls['start_time']}"

            elif class_start <= now <= activation_deadline:
                cls_dict["status_type"] = "CAN_START"
                cls_dict["status_text"] = "Can Start"

            elif activation_deadline < now <= class_end:
                cls_dict["status_type"] = "NOT_HELD"
                cls_dict["status_text"] = "Class Not Held"

            else:
                cls_dict["status_type"] = "TIME_OVER"
                cls_dict["status_text"] = "Time Over"

        subjects.append(cls_dict)

    con.close()

    return render_template(
        "teacher_section.html",
        section=section,
        today=today,
        subjects=subjects
    )


@app.route("/teacher-subject/<section>/<subject>")
def teacher_subject(section, subject):
    if "teacher_db_id" not in session:
        return redirect(url_for("teacher_login"))

    close_expired_active_sessions()

    section = normalize_section(section)
    subject = subject.strip().upper()

    con = get_db()

    sessions = con.execute("""
        SELECT *
        FROM class_sessions
        WHERE teacher_id = ?
        AND UPPER(REPLACE(TRIM(section), ' ', '-')) = ?
        AND UPPER(subject) = ?
        ORDER BY class_date DESC, id DESC
    """, (
        session["teacher_db_id"],
        section,
        subject
    )).fetchall()

    con.close()

    return render_template(
        "teacher_subject.html",
        section=section,
        subject=subject,
        sessions=sessions
    )


@app.route("/start-class/<int:timetable_id>", methods=["POST"])
def start_class(timetable_id):
    if "teacher_db_id" not in session:
        return redirect(url_for("teacher_login"))

    teacher_lat = request.form.get("teacher_lat")
    teacher_lng = request.form.get("teacher_lng")

    if not teacher_lat or not teacher_lng:
        return "Location permission required. Teacher location is needed to start class."

    con = get_db()

    class_data = con.execute("""
        SELECT timetable.*, rooms.room_lat, rooms.room_lng, rooms.room_range
        FROM timetable
        LEFT JOIN rooms ON timetable.room_no = rooms.room_no
        WHERE timetable.id = ? AND teacher_id = ?
    """, (timetable_id, session["teacher_db_id"])).fetchone()

    if not class_data:
        con.close()
        return "Class not found"

    now = datetime.now()
    class_date = now.strftime("%Y-%m-%d")

    class_start = datetime.strptime(
        class_date + " " + class_data["start_time"],
        "%Y-%m-%d %H:%M"
    )

    class_end = datetime.strptime(
        class_date + " " + class_data["end_time"],
        "%Y-%m-%d %H:%M"
    )

    activation_deadline = class_start + timedelta(minutes=TEACHER_ACTIVATION_LIMIT_MINUTES)
    class_time = f"{class_data['start_time']} - {class_data['end_time']}"
    normalized_section = normalize_section(class_data["section"])

    existing_session = con.execute("""
        SELECT * FROM class_sessions
        WHERE teacher_id = ?
        AND UPPER(subject) = ?
        AND UPPER(REPLACE(TRIM(section), ' ', '-')) = ?
        AND class_date = ?
        AND class_time = ?
        ORDER BY id DESC LIMIT 1
    """, (
        session["teacher_db_id"],
        class_data["subject"].upper(),
        normalized_section,
        class_date,
        class_time
    )).fetchone()

    if existing_session:
        con.close()

        if existing_session["status"] == "ACTIVE":
            return "This class is already active today. Duplicate class session is not allowed."

        if existing_session["status"] == "CLOSED":
            return "This class attendance is already completed today. Duplicate record is not allowed."

        if existing_session["status"] == "NOT_HELD":
            return "This class was already marked as Not Held today. Duplicate record is not allowed."

        return "This class record already exists today. Duplicate record is not allowed."

    if now < class_start:
        con.close()
        return f"Class abhi start nahi hui hai. Start time: {class_data['start_time']}"

    if now > activation_deadline:
        con.execute("""
            INSERT INTO class_sessions
            (teacher_id, subject, section, class_date, class_time, activated_at,
             attendance_end, session_code, status, timetable_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            session["teacher_db_id"],
            class_data["subject"],
            normalized_section,
            class_date,
            class_time,
            now.strftime("%H:%M:%S"),
            "-",
            "-",
            "NOT_HELD",
            timetable_id
        ))

        con.commit()
        con.close()

        return "Class Not Held. Teacher activation time over. Students ki attendance affect nahi hogi."

    if now > class_end:
        con.close()
        return "Class time already over."

    if class_data["room_lat"] is None or class_data["room_lng"] is None:
        con.close()
        return "Room location not set. Please ask Admin to set the GPS location for this room."

    teacher_distance = calculate_distance(
        float(teacher_lat),
        float(teacher_lng),
        float(class_data["room_lat"]),
        float(class_data["room_lng"])
    )

    allowed_range = class_data["room_range"] if class_data["room_range"] else TEACHER_ALLOWED_RADIUS_METERS

    if teacher_distance > allowed_range:
        con.close()
        return f"Blocked: Teacher is outside the room range. Distance: {int(teacher_distance)} meters (Allowed: {allowed_range}m)"

    activated_at = now.strftime("%H:%M:%S")
    attendance_end = (now + timedelta(minutes=ATTENDANCE_WINDOW_MINUTES)).strftime("%H:%M:%S")
    session_code = generate_code()

    con.execute("""
        UPDATE class_sessions
        SET status = 'CLOSED'
        WHERE teacher_id = ? AND class_date = ? AND status = 'ACTIVE'
    """, (session["teacher_db_id"], class_date))

    cur = con.execute("""
        INSERT INTO class_sessions
        (teacher_id, subject, section, class_date, class_time, activated_at,
         attendance_end, session_code, status, timetable_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        session["teacher_db_id"],
        class_data["subject"],
        normalized_section,
        class_date,
        class_time,
        activated_at,
        attendance_end,
        session_code,
        "ACTIVE",
        timetable_id
    ))

    session_id = cur.lastrowid

    students = con.execute(
        "SELECT * FROM students WHERE UPPER(REPLACE(TRIM(section), ' ', '-')) = ?",
        (normalized_section,)
    ).fetchall()

    for student in students:
        con.execute("""
            INSERT OR IGNORE INTO attendance(session_id, student_id, date, time, status)
            VALUES (?, ?, ?, NULL, 'ABSENT')
        """, (session_id, student["id"], class_date))

    con.commit()
    con.close()

    return redirect(url_for("teacher_section", section=normalized_section))


@app.route("/close-class/<int:session_id>", methods=["POST"])
def close_class(session_id):
    if "teacher_db_id" not in session:
        return redirect(url_for("teacher_login"))

    con = get_db()

    active_session = con.execute("""
        SELECT * FROM class_sessions
        WHERE id = ? AND teacher_id = ?
    """, (session_id, session["teacher_db_id"])).fetchone()

    con.execute("""
        UPDATE class_sessions
        SET status = 'CLOSED'
        WHERE id = ? AND teacher_id = ?
    """, (session_id, session["teacher_db_id"]))

    con.commit()
    con.close()

    if active_session:
        return redirect(url_for("teacher_section", section=active_session["section"]))

    return redirect(url_for("teacher_dashboard"))


@app.route("/student-dashboard")
def student_dashboard():
    if "student_db_id" not in session:
        return redirect(url_for("student_login"))

    close_expired_active_sessions()

    current_date = datetime.now().strftime("%Y-%m-%d")

    con = get_db()

    active_session = con.execute("""
        SELECT * FROM class_sessions
        WHERE class_date = ?
        AND UPPER(REPLACE(TRIM(section), ' ', '-')) = ?
        AND status = 'ACTIVE'
        ORDER BY id DESC LIMIT 1
    """, (current_date, normalize_section(session["section"]))).fetchone()

    attendance_history = con.execute("""
        SELECT 
            class_sessions.subject,
            class_sessions.section,
            class_sessions.class_date,
            class_sessions.class_time,
            class_sessions.status AS class_status,
            attendance.time,
            attendance.status
        FROM attendance
        JOIN class_sessions ON class_sessions.id = attendance.session_id
        WHERE attendance.student_id = ?
        AND NOT (class_sessions.status = 'ACTIVE' AND attendance.status = 'ABSENT')
        ORDER BY class_sessions.subject ASC, class_sessions.class_date DESC, attendance.id DESC
    """, (session["student_db_id"],)).fetchall()

    total_classes = len(attendance_history)
    present_count = 0
    absent_count = 0

    for row in attendance_history:
        if row["status"] == "PRESENT":
            present_count += 1
        elif row["status"] == "ABSENT":
            absent_count += 1

    if total_classes > 0:
        attendance_percentage = round((present_count / total_classes) * 100, 2)
    else:
        attendance_percentage = 0

    subject_groups = {}

    for row in attendance_history:
        subject = row["subject"]

        if subject not in subject_groups:
            subject_groups[subject] = {
                "total": 0,
                "present": 0,
                "absent": 0,
                "percentage": 0,
                "records": []
            }

        subject_groups[subject]["records"].append(row)
        subject_groups[subject]["total"] += 1

        if row["status"] == "PRESENT":
            subject_groups[subject]["present"] += 1
        elif row["status"] == "ABSENT":
            subject_groups[subject]["absent"] += 1

    for subject in subject_groups:
        total = subject_groups[subject]["total"]
        present = subject_groups[subject]["present"]

        if total > 0:
            subject_groups[subject]["percentage"] = round((present / total) * 100, 2)

    current_day = datetime.now().strftime("%A")
    today_timetable = con.execute("""
        SELECT timetable.*, teachers.name as teacher_name
        FROM timetable 
        JOIN teachers ON timetable.teacher_id = teachers.id
        WHERE day = ? AND section = ?
        ORDER BY start_time ASC
    """, (current_day, normalize_section(session["section"]))).fetchall()

    con.close()

    return render_template(
        "student_dashboard.html",
        active_session=active_session,
        total_classes=total_classes,
        present_count=present_count,
        absent_count=absent_count,
        attendance_percentage=attendance_percentage,
        subject_groups=subject_groups,
        today_timetable=today_timetable
    )


@app.route("/scan/<int:session_id>")
def scan_attendance(session_id):
    if "student_db_id" not in session:
        return redirect(url_for("student_login"))

    close_expired_active_sessions()

    con = get_db()

    active_session = con.execute("""
        SELECT * FROM class_sessions
        WHERE id = ? AND status = 'ACTIVE'
    """, (session_id,)).fetchone()

    con.close()

    if not active_session:
        return redirect(url_for("student_dashboard"))

    return render_template("scan.html", active_session=active_session)


@app.route("/api/mark-attendance", methods=["POST"])
def api_mark_attendance():
    if "student_db_id" not in session:
        return jsonify(success=False, message="Student login required")

    close_expired_active_sessions()

    data = request.get_json()

    session_id = int(data.get("session_id"))
    entered_code = data.get("session_code", "").strip().upper()
    image_data = data.get("image_data")
    lat = data.get("lat")
    lng = data.get("lng")
    blink_verified = data.get("blink_verified", False)

    if not entered_code:
        return jsonify(success=False, message="Session code required")

    if not image_data:
        return jsonify(success=False, message="Camera image required")

    if not blink_verified:
        return jsonify(success=False, message="Blink verification required")

    if lat is None or lng is None:
        return jsonify(success=False, message="Location permission required")

    con = get_db()

    active_session = con.execute("""
        SELECT class_sessions.*, rooms.room_lat, rooms.room_lng, rooms.room_range
        FROM class_sessions
        LEFT JOIN timetable ON class_sessions.timetable_id = timetable.id
        LEFT JOIN rooms ON timetable.room_no = rooms.room_no
        WHERE class_sessions.id = ? AND class_sessions.status = 'ACTIVE'
    """, (session_id,)).fetchone()

    if not active_session:
        con.close()
        return jsonify(success=False, message="Attendance window closed")

    if entered_code != active_session["session_code"].upper():
        con.close()
        return jsonify(success=False, message="Wrong session code")

    now = datetime.now()

    attendance_end_time = datetime.strptime(
        active_session["class_date"] + " " + active_session["attendance_end"],
        "%Y-%m-%d %H:%M:%S"
    )

    if now > attendance_end_time:
        con.execute("""
            UPDATE class_sessions
            SET status = 'CLOSED'
            WHERE id = ?
        """, (session_id,))

        con.commit()
        con.close()

        return jsonify(success=False, message="Attendance window closed")

    student = con.execute("""
        SELECT * FROM students
        WHERE id = ?
    """, (session["student_db_id"],)).fetchone()

    if not student:
        con.close()
        return jsonify(success=False, message="Student record not found")

    already_present = con.execute("""
        SELECT * FROM attendance
        WHERE session_id = ? AND student_id = ? AND status = 'PRESENT'
    """, (session_id, session["student_db_id"])).fetchone()

    if already_present:
        con.close()
        return jsonify(success=False, message="Already verified for this class")

    face_ok, face_message = verify_student_face(image_data, student["roll_no"])

    if not face_ok:
        con.close()
        return jsonify(
            success=False,
            message="Face verification failed: " + face_message
        )

    if active_session["room_lat"] is None or active_session["room_lng"] is None:
        con.close()
        return jsonify(success=False, message="Room location is not configured. Admin needs to set GPS for this room.")

    distance = calculate_distance(
        float(lat),
        float(lng),
        float(active_session["room_lat"]),
        float(active_session["room_lng"])
    )

    allowed_range = active_session["room_range"] if active_session["room_range"] else ALLOWED_RADIUS_METERS

    if distance > allowed_range:
        con.close()
        return jsonify(
            success=False,
            message=f"Blocked: You are outside allowed radius. Distance: {int(distance)} meters (Allowed: {allowed_range}m)"
        )

    proof_filename = save_proof_image(image_data, student["roll_no"], session_id)
    proof_url = url_for("proof_image", filename=proof_filename)

    existing_attendance = con.execute("""
        SELECT * FROM attendance
        WHERE session_id = ? AND student_id = ?
    """, (session_id, session["student_db_id"])).fetchone()

    if existing_attendance:
        con.execute("""
            UPDATE attendance
            SET status = 'PRESENT',
                time = ?,
                distance_meters = ?,
                proof_image = ?
            WHERE session_id = ? AND student_id = ?
        """, (
            now.strftime("%H:%M:%S"),
            round(distance, 2),
            proof_filename,
            session_id,
            session["student_db_id"]
        ))
    else:
        con.execute("""
            INSERT INTO attendance(session_id, student_id, date, time, status, distance_meters, proof_image)
            VALUES (?, ?, ?, ?, 'PRESENT', ?, ?)
        """, (
            session_id,
            session["student_db_id"],
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            round(distance, 2),
            proof_filename
        ))

    con.commit()
    con.close()

    return jsonify(
        success=True,
        message="Attendance marked successfully",
        name=student["name"],
        roll_no=student["roll_no"],
        subject=active_session["subject"],
        distance=round(distance, 2),
        proof_image=proof_url,
        face_status=face_message,
        blink_status="Verified"
    )


@app.route("/report/<int:session_id>")
def report(session_id):
    if "teacher_db_id" not in session:
        return redirect(url_for("teacher_login"))

    con = get_db()

    active_session = con.execute("""
        SELECT * FROM class_sessions
        WHERE id = ? AND teacher_id = ?
    """, (session_id, session["teacher_db_id"])).fetchone()

    if not active_session:
        con.close()
        return "Report not found"

    attendance_rows = con.execute("""
        SELECT 
            students.id AS student_id,
            students.name,
            students.roll_no,
            COALESCE(attendance.time, '-') AS time,
            COALESCE(attendance.status, 'ABSENT') AS status,
            COALESCE(attendance.distance_meters, '-') AS distance_meters,
            COALESCE(attendance.proof_image, '-') AS proof_image
        FROM students
        LEFT JOIN attendance 
            ON attendance.student_id = students.id 
            AND attendance.session_id = ?
        WHERE UPPER(REPLACE(TRIM(students.section), ' ', '-')) = ?
        ORDER BY students.name
    """, (session_id, normalize_section(active_session["section"]))).fetchall()

    con.close()

    return render_template(
        "report.html",
        active_session=active_session,
        attendance_rows=attendance_rows
    )


@app.route("/export-excel/<int:session_id>")
def export_excel(session_id):
    if "teacher_db_id" not in session:
        return redirect(url_for("teacher_login"))

    con = get_db()

    active_session = con.execute("""
        SELECT * FROM class_sessions
        WHERE id = ? AND teacher_id = ?
    """, (session_id, session["teacher_db_id"])).fetchone()

    if not active_session:
        con.close()
        return "Report not found"

    attendance_rows = con.execute("""
        SELECT 
            students.name,
            students.roll_no,
            COALESCE(attendance.time, '-') AS time,
            COALESCE(attendance.status, 'ABSENT') AS status,
            COALESCE(attendance.distance_meters, '-') AS distance_meters,
            COALESCE(attendance.proof_image, '-') AS proof_image
        FROM students
        LEFT JOIN attendance 
            ON attendance.student_id = students.id 
            AND attendance.session_id = ?
        WHERE UPPER(REPLACE(TRIM(students.section), ' ', '-')) = ?
        ORDER BY students.name
    """, (session_id, normalize_section(active_session["section"]))).fetchall()

    con.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Smart Attendance Report - {active_session['subject']}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="102A43")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["Subject", active_session["subject"]])
    ws.append(["Section", active_session["section"]])
    ws.append(["Date", active_session["class_date"]])
    ws.append(["Class Time", active_session["class_time"]])
    ws.append(["Status", active_session["status"]])
    ws.append([])

    headers = [
        "Name",
        "Roll No",
        "Subject",
        "Section",
        "Date",
        "Class Time",
        "Attendance Time",
        "Status",
        "Distance (meters)",
        "Proof Image"
    ]

    ws.append(headers)

    header_row = ws.max_row
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in attendance_rows:
        ws.append([
            row["name"],
            row["roll_no"],
            active_session["subject"],
            active_session["section"],
            active_session["class_date"],
            active_session["class_time"],
            row["time"],
            row["status"],
            row["distance_meters"],
            row["proof_image"]
        ])

    column_widths = {
        "A": 24,
        "B": 18,
        "C": 18,
        "D": 14,
        "E": 16,
        "F": 20,
        "G": 18,
        "H": 14,
        "I": 20,
        "J": 45
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    safe_subject = str(active_session["subject"]).replace(" ", "_")
    filename = f"Attendance_{safe_subject}_{active_session['class_date']}_Session_{session_id}.xlsx"
    filepath = os.path.join(EXPORT_FOLDER, filename)

    wb.save(filepath)

    return send_file(filepath, as_attachment=True)


@app.route("/admin-login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        admin_id = request.form.get("admin_id", "").strip()
        password = request.form.get("password", "").strip()

        con = get_db()
        admin = con.execute(
            "SELECT * FROM admins WHERE admin_id = ? AND password = ?",
            (admin_id, password)
        ).fetchone()
        con.close()

        if admin:
            session.clear()
            session["admin_id"] = admin["id"]
            session["admin_name"] = admin["admin_id"]
            return redirect(url_for("admin_dashboard"))

        return render_template("admin_login.html", error="Wrong Admin ID or Password")

    return render_template("admin_login.html")


@app.route("/admin-dashboard")
def admin_dashboard():
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    con = get_db()

    total_students = con.execute("SELECT COUNT(*) AS total FROM students").fetchone()["total"]
    total_teachers = con.execute("SELECT COUNT(*) AS total FROM teachers").fetchone()["total"]
    total_sessions = con.execute("SELECT COUNT(*) AS total FROM class_sessions").fetchone()["total"]
    total_timetable = con.execute("SELECT COUNT(*) AS total FROM timetable").fetchone()["total"]

    teachers = con.execute("""
        SELECT * FROM teachers
        ORDER BY name
    """).fetchall()

    all_sections = con.execute("""
        SELECT section FROM students
        UNION
        SELECT section FROM timetable
        ORDER BY section
    """).fetchall()

    course_groups = {}

    explicit_courses = con.execute("SELECT course_name FROM courses").fetchall()
    for row in explicit_courses:
        course = row["course_name"].upper()
        if course not in course_groups:
            course_groups[course] = {
                "course": course,
                "sections": set(),
                "student_count": 0,
                "timetable_count": 0
            }

    explicit_sections = con.execute("SELECT course_name, section_name FROM sections").fetchall()
    for row in explicit_sections:
        course = row["course_name"].upper()
        section_full = f"{course}-{row['section_name'].upper()}"
        
        if course not in course_groups:
            course_groups[course] = {
                "course": course,
                "sections": set(),
                "student_count": 0,
                "timetable_count": 0
            }
        course_groups[course]["sections"].add(section_full)

    for row in all_sections:
        section = normalize_section(row["section"])
        course = get_course_from_section(section)

        if course not in course_groups:
            course_groups[course] = {
                "course": course,
                "sections": set(),
                "student_count": 0,
                "timetable_count": 0
            }

        course_groups[course]["sections"].add(section)

    for course in course_groups:
        like_pattern = course + "-%"

        student_count = con.execute("""
            SELECT COUNT(*) AS total
            FROM students
            WHERE UPPER(REPLACE(TRIM(section), ' ', '-')) LIKE ?
        """, (like_pattern,)).fetchone()["total"]

        timetable_count = con.execute("""
            SELECT COUNT(*) AS total
            FROM timetable
            WHERE UPPER(REPLACE(TRIM(section), ' ', '-')) LIKE ?
        """, (like_pattern,)).fetchone()["total"]

        course_groups[course]["student_count"] = student_count
        course_groups[course]["timetable_count"] = timetable_count
        course_groups[course]["sections"] = sorted(list(course_groups[course]["sections"]))

    con.close()

    return render_template(
        "admin_dashboard.html",
        total_students=total_students,
        total_teachers=total_teachers,
        total_sessions=total_sessions,
        total_timetable=total_timetable,
        teachers=teachers,
        course_groups=course_groups
    )


@app.route("/admin-course/<course>")
def admin_course(course):
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    course = course.strip().upper()
    like_pattern = course + "-%"

    con = get_db()

    timetable_rows = con.execute("""
        SELECT 
            timetable.id,
            timetable.day,
            timetable.subject,
            timetable.start_time,
            timetable.end_time,
            timetable.section,
            teachers.name AS teacher_name,
            teachers.teacher_id AS teacher_code
        FROM timetable
        JOIN teachers ON teachers.id = timetable.teacher_id
        WHERE UPPER(REPLACE(TRIM(timetable.section), ' ', '-')) LIKE ? OR UPPER(TRIM(timetable.section)) = ?
        ORDER BY 
            timetable.section,
            CASE timetable.day
                WHEN 'Monday' THEN 1
                WHEN 'Tuesday' THEN 2
                WHEN 'Wednesday' THEN 3
                WHEN 'Thursday' THEN 4
                WHEN 'Friday' THEN 5
                WHEN 'Saturday' THEN 6
                WHEN 'Sunday' THEN 7
                ELSE 8
            END,
            timetable.start_time
    """, (like_pattern, course)).fetchall()

    all_sections = get_all_sections_for_course(con, course)
    timetable_groups = {sec: [] for sec in all_sections}

    for row in timetable_rows:
        section = normalize_section(row["section"])
        if section not in timetable_groups:
            timetable_groups[section] = []
        timetable_groups[section].append(row)

    student_rows = con.execute("""
        SELECT *
        FROM students
        WHERE UPPER(REPLACE(TRIM(section), ' ', '-')) LIKE ? OR UPPER(TRIM(section)) = ?
        ORDER BY section, name
    """, (like_pattern, course)).fetchall()

    student_groups = {sec: [] for sec in all_sections}

    for row in student_rows:
        section = normalize_section(row["section"])
        if section not in student_groups:
            student_groups[section] = []
        student_groups[section].append(row)

    con.close()

    return render_template(
        "admin_course.html",
        course=course,
        timetable_groups=timetable_groups,
        student_groups=student_groups,
        timetable_rows=timetable_rows,
        student_rows=student_rows
    )


@app.route("/admin-course/<course>/students")
def admin_course_students(course):
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    course = course.strip().upper()
    like_pattern = course + "-%"

    con = get_db()

    student_rows = con.execute("""
        SELECT *
        FROM students
        WHERE UPPER(REPLACE(TRIM(section), ' ', '-')) LIKE ? OR UPPER(TRIM(section)) = ?
        ORDER BY section, name
    """, (like_pattern, course)).fetchall()

    all_sections = get_all_sections_for_course(con, course)
    student_groups = {sec: [] for sec in all_sections}

    for row in student_rows:
        section = normalize_section(row["section"])
        if section not in student_groups:
            student_groups[section] = []
        student_groups[section].append(row)

    con.close()

    return render_template(
        "admin_course_students.html",
        course=course,
        student_groups=student_groups,
        student_rows=student_rows
    )


@app.route("/admin/add-teacher", methods=["GET", "POST"])
def add_teacher():
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        teacher_id = request.form.get("teacher_id", "").strip()
        password = request.form.get("password", "").strip()

        if not name or not teacher_id or not password:
            return render_template("add_teacher.html", error="All fields are required")

        con = get_db()

        existing = con.execute(
            "SELECT * FROM teachers WHERE teacher_id = ?",
            (teacher_id,)
        ).fetchone()

        if existing:
            con.close()
            return render_template("add_teacher.html", error="Teacher ID already exists")

        con.execute("""
            INSERT INTO teachers(name, teacher_id, password)
            VALUES (?, ?, ?)
        """, (name, teacher_id, password))

        con.commit()
        con.close()

        return redirect(url_for("admin_dashboard"))

    return render_template("add_teacher.html")


@app.route("/admin/delete-teacher/<int:teacher_db_id>", methods=["POST"])
def delete_teacher(teacher_db_id):
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    con = get_db()

    teacher = con.execute(
        "SELECT * FROM teachers WHERE id = ?",
        (teacher_db_id,)
    ).fetchone()

    if not teacher:
        con.close()
        return redirect(url_for("admin_dashboard"))

    con.execute(
        "DELETE FROM timetable WHERE teacher_id = ?",
        (teacher_db_id,)
    )

    con.execute(
        "DELETE FROM teachers WHERE id = ?",
        (teacher_db_id,)
    )

    con.commit()
    con.close()

    return redirect(url_for("admin_dashboard"))


@app.route("/admin/add-timetable", methods=["GET", "POST"])
def add_timetable():
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    con = get_db()

    teachers = con.execute("""
        SELECT * FROM teachers
        ORDER BY name
    """).fetchall()

    rooms = con.execute("SELECT * FROM rooms ORDER BY room_no").fetchall()
    sections = get_all_sections(con)

    if request.method == "POST":
        day = request.form.get("day", "").strip()
        subject = request.form.get("subject", "").strip().upper()
        start_time = request.form.get("start_time", "").strip()
        end_time = request.form.get("end_time", "").strip()
        section = normalize_section(request.form.get("section", ""))
        teacher_id = request.form.get("teacher_id", "").strip()
        room_no = request.form.get("room_no", "").strip().upper()

        if not day or not subject or not start_time or not end_time or not section or not teacher_id or not room_no:
            con.close()
            return render_template(
                "add_timetable.html",
                teachers=teachers,
                rooms=rooms,
                sections=sections,
                error="All fields are required"
            )

        if start_time >= end_time:
            con.close()
            return render_template(
                "add_timetable.html",
                teachers=teachers,
                rooms=rooms,
                sections=sections,
                error="End time must be greater than start time"
            )

        teacher = con.execute(
            "SELECT * FROM teachers WHERE id = ?",
            (teacher_id,)
        ).fetchone()

        if not teacher:
            con.close()
            return render_template(
                "add_timetable.html",
                teachers=teachers,
                rooms=rooms,
                sections=sections,
                error="Invalid teacher selected"
            )

        existing = con.execute("""
            SELECT * FROM timetable
            WHERE day = ? AND subject = ? AND start_time = ? AND section = ? AND teacher_id = ?
        """, (day, subject, start_time, section, teacher_id)).fetchone()

        if existing:
            con.close()
            return render_template(
                "add_timetable.html",
                teachers=teachers,
                rooms=rooms,
                sections=sections,
                error="This timetable entry already exists"
            )

        room_data = con.execute("SELECT room_range FROM rooms WHERE UPPER(room_no) = ?", (room_no.upper(),)).fetchone()
        if not room_data:
            con.close()
            return render_template(
                "add_timetable.html",
                teachers=teachers,
                rooms=rooms,
                sections=sections,
                error="Selected room does not exist"
            )
        room_range_val = room_data["room_range"]

        con.execute("""
            INSERT INTO timetable(day, subject, start_time, end_time, section, teacher_id, room_no, room_range)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (day, subject, start_time, end_time, section, teacher_id, room_no, room_range_val))

        con.commit()
        con.close()

        return redirect(url_for("admin_dashboard"))

    con.close()
    return render_template("add_timetable.html", teachers=teachers, rooms=rooms, sections=sections)

@app.route("/admin/add-room", methods=["GET", "POST"])
def add_room():
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    if request.method == "POST":
        room_no = request.form.get("room_no", "").strip().upper()
        room_range = request.form.get("room_range", "").strip()
        room_lat = request.form.get("room_lat", "").strip()
        room_lng = request.form.get("room_lng", "").strip()

        if not room_no or not room_range or not room_lat or not room_lng:
            return render_template("add_room.html", error="All fields including Location are required")
            
        try:
            room_range_val = int(room_range)
            room_lat_val = float(room_lat)
            room_lng_val = float(room_lng)
        except ValueError:
            return render_template("add_room.html", error="Range, Latitude and Longitude must be valid numbers")
            
        con = get_db()
        try:
            con.execute("INSERT INTO rooms (room_no, room_range, room_lat, room_lng) VALUES (?, ?, ?, ?)", (room_no, room_range_val, room_lat_val, room_lng_val))
            con.commit()
        except sqlite3.IntegrityError:
            con.close()
            return render_template("add_room.html", error="Room already exists")
            
        con.close()
        return redirect(url_for("admin_dashboard"))

    return render_template("add_room.html")


@app.route("/admin/delete-timetable/<int:timetable_id>", methods=["POST"])
def delete_timetable(timetable_id):
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    con = get_db()

    con.execute(
        "DELETE FROM timetable WHERE id = ?",
        (timetable_id,)
    )

    con.commit()
    con.close()

    return redirect(url_for("admin_dashboard"))

@app.route("/admin/add-course", methods=["GET", "POST"])
def add_course():
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    if request.method == "POST":
        course_name = request.form.get("course_name", "").strip().upper()
        if not course_name:
            return render_template("add_course.html", error="Course Name is required")
            
        con = get_db()
        try:
            con.execute("INSERT INTO courses (course_name) VALUES (?)", (course_name,))
            con.commit()
        except sqlite3.IntegrityError:
            con.close()
            return render_template("add_course.html", error="Course already exists")
            
        con.close()
        return redirect(url_for("admin_dashboard"))

    return render_template("add_course.html")


@app.route("/admin/add-section", methods=["GET", "POST"])
def add_section():
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    con = get_db()
    courses = con.execute("SELECT course_name FROM courses ORDER BY course_name").fetchall()
    
    if request.method == "POST":
        course_name = request.form.get("course_name", "").strip().upper()
        section_name = request.form.get("section_name", "").strip().upper()
        
        if not course_name or not section_name:
            con.close()
            return render_template("add_section.html", courses=courses, error="Course and Section Name are required")
            
        try:
            con.execute("INSERT INTO sections (course_name, section_name) VALUES (?, ?)", (course_name, section_name))
            con.commit()
        except sqlite3.IntegrityError:
            con.close()
            return render_template("add_section.html", courses=courses, error="Section already exists for this course")
            
        con.close()
        return redirect(url_for("admin_dashboard"))

    con.close()
    return render_template("add_section.html", courses=courses)


@app.route("/admin/add-student", methods=["GET", "POST"])
def add_student():
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    selected_course = request.args.get("course", "").strip().upper()
    con = get_db()
    all_sections = get_all_sections(con)
    all_courses = [r["course_name"] for r in con.execute("SELECT course_name FROM courses ORDER BY course_name").fetchall()]

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        roll_no = request.form.get("roll_no", "").strip().upper()
        password = request.form.get("password", "").strip()
        raw_sec = request.form.get("section", "").strip().upper()
        
        if selected_course and not raw_sec.startswith(selected_course + "-") and "-" not in raw_sec:
            raw_sec = f"{selected_course}-{raw_sec}"

        section = normalize_section(raw_sec)
        photo = request.files.get("photo")

        if not name or not roll_no or not password or not section:
            con.close()
            return render_template("add_student.html", error="All fields are required", sections=all_sections, courses=all_courses, selected_course=selected_course)

        if not photo or photo.filename == "":
            con.close()
            return render_template("add_student.html", error="Student photo is required", sections=all_sections, courses=all_courses, selected_course=selected_course)

        allowed_extensions = [".jpg", ".jpeg", ".png", ".webp"]
        original_filename = secure_filename(photo.filename)
        ext = os.path.splitext(original_filename)[1].lower()

        if ext not in allowed_extensions:
            con.close()
            return render_template("add_student.html", error="Only JPG, JPEG, PNG, WEBP images are allowed", sections=all_sections, courses=all_courses, selected_course=selected_course)

        saved_filename = roll_no + ext
        saved_path = os.path.join(STUDENT_IMAGE_FOLDER, saved_filename)

        existing = con.execute(
            "SELECT * FROM students WHERE UPPER(roll_no) = ?",
            (roll_no,)
        ).fetchone()

        if existing:
            con.close()
            return render_template("add_student.html", error="Student with this roll number already exists", sections=all_sections, courses=all_courses, selected_course=selected_course)

        photo.save(saved_path)

        con.execute("""
            INSERT INTO students(name, roll_no, password, section)
            VALUES (?, ?, ?, ?)
        """, (name, roll_no, password, section))

        con.commit()
        con.close()

        if selected_course:
            return redirect(f"/admin-course/{selected_course}/students")
        return redirect(url_for("admin_dashboard"))

    con.close()
    return render_template("add_student.html", sections=all_sections, courses=all_courses, selected_course=selected_course)


@app.route("/admin/delete-student/<int:student_id>", methods=["POST"])
def delete_student(student_id):
    if "admin_id" not in session:
        return redirect(url_for("admin_login"))

    con = get_db()

    student = con.execute(
        "SELECT * FROM students WHERE id = ?",
        (student_id,)
    ).fetchone()

    if student:
        roll_no = student["roll_no"].strip().upper()

        for ext in [".jpg", ".jpeg", ".png", ".webp"]:
            image_path = os.path.join(STUDENT_IMAGE_FOLDER, roll_no + ext)
            if os.path.exists(image_path):
                os.remove(image_path)

        con.execute("DELETE FROM attendance WHERE student_id = ?", (student_id,))
        con.execute("DELETE FROM students WHERE id = ?", (student_id,))
        con.commit()

    con.close()

    return redirect(url_for("admin_dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
